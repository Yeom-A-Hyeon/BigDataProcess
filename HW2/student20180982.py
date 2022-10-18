#!/usr/bin/python3

from openpyxl import load_workbook
import math

wb = load_workbook(filename = 'student.xlsx')
ws = wb['Sheet1']

li = []

row_id = 0
for row in ws:
    row_id += 1
    if row_id != 1:
        total = 0
        total += ws.cell(row=row_id, column=3).value * 0.30
        total += ws.cell(row=row_id, column=4).value * 0.35
        total += ws.cell(row=row_id, column=5).value * 0.34
        total += ws.cell(row=row_id, column=6).value
        ws.cell(row=row_id, column=7).value = total
        li.append((row_id, total))

li.sort(key=lambda x : x[1], reverse=True)

a_cnt = math.floor(len(li) * 0.3)
aplus_cnt = math.floor(a_cnt * 0.5)
b_cnt = math.floor(len(li) * 0.7) - a_cnt
bplus_cnt = math.floor(b_cnt * 0.5)
c_cnt = len(li) - a_cnt - b_cnt
cplus_cnt = math.floor(c_cnt * 0.5)

cnt = 0
for i in range(aplus_cnt):
    ws.cell(row=li[cnt][0], column=8).value = 'A+'
    cnt += 1
for i in range(a_cnt-aplus_cnt):
    ws.cell(row=li[cnt][0], column=8).value = 'A0'
    cnt += 1
for i in range(bplus_cnt):
    ws.cell(row=li[cnt][0], column=8).value = 'B+'
    cnt += 1
for i in range(b_cnt-bplus_cnt):
    ws.cell(row=li[cnt][0], column=8).value = 'B0'
    cnt += 1
for i in range(cplus_cnt):
    ws.cell(row=li[cnt][0], column=8).value = 'C+'
    cnt += 1
for i in range(c_cnt-cplus_cnt):
    ws.cell(row=li[cnt][0], column=8).value = 'C0'
    cnt += 1

wb.save('student.xlsx')